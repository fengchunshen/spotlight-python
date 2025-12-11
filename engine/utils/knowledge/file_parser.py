from __future__ import annotations

import os
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4
from urllib.parse import urlparse

import aiofiles
import httpx
import pandas as pd
from langchain_text_splitters import RecursiveCharacterTextSplitter
from markdownify import markdownify as md
from pypdf import PdfReader

from engine.logging_utils import get_logger, sanitize_log_message
from engine.utils.storage.upload_client import UploadClient

SUPPORTED_FILE_EXTENSIONS: tuple[str, ...] = (
    ".txt",
    ".md",
    ".docx",
    ".doc",
    ".html",
    ".htm",
    ".csv",
    ".xls",
    ".xlsx",
    ".pdf",
    ".jpg",
    ".jpeg",
    ".png",
    ".bmp",
    ".tiff",
    ".tif",
)


def is_supported_file_extension(file_name: str | os.PathLike[str]) -> bool:
    """检查文件是否为支持的扩展名"""
    return Path(file_name).suffix.lower() in SUPPORTED_FILE_EXTENSIONS


def _build_text_splitter(chunk_size: int, chunk_overlap: int) -> RecursiveCharacterTextSplitter:
    return RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ".", " ", ""],
    )


def chunk_text(text: str, chunk_size: int = 500, chunk_overlap: int = 100) -> list[dict[str, Any]]:
    """
    将文本切分为固定大小的块

    :param text: 待切分文本
    :param chunk_size: 块大小
    :param chunk_overlap: 重叠大小
    :return:
    """
    splitter = _build_text_splitter(chunk_size, chunk_overlap)
    nodes = splitter.split_text(text)
    return [{"text": node, "metadata": {"chunk_idx": idx}} for idx, node in enumerate(nodes)]


async def chunk_file(
    file_path: str,
    chunk_size: int = 500,
    chunk_overlap: int = 100,
    trace_id: str | None = None,
) -> list[dict[str, Any]]:
    """
    读取文件并切分

    :param file_path: 文件路径
    :param chunk_size: 块大小
    :param chunk_overlap: 重叠大小
    :param trace_id: 链路追踪 ID
    :return:
    """
    logger = get_logger(trace_id)
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    content = await _read_text_file(path, trace_id)
    if not content:
        logger.warning("文件内容为空，跳过切分")
        return []

    return chunk_text(content, chunk_size=chunk_size, chunk_overlap=chunk_overlap)


async def process_file_to_markdown(
    file_path: str,
    params: dict[str, Any] | None = None,
    uploader: UploadClient | None = None,
    trace_id: str | None = None,
) -> str:
    """
    将文件转换为 Markdown 文本，支持 MinIO URL

    :param file_path: 本地路径或 MinIO URL
    :param params: 处理参数，如 chunk_size/db_id/enable_ocr
    :param uploader: 上传客户端，用于 docx 内嵌图片
    :param trace_id: 链路追踪 ID
    :return:
    """
    params = params or {}
    logger = get_logger(trace_id)

    actual_path: Path
    cleanup_path: Path | None = None

    if _is_http_url(file_path):
        cleanup_path = Path(tempfile.mkstemp(prefix="kb_", suffix=Path(file_path).suffix)[1])
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(file_path)
                resp.raise_for_status()
                async with aiofiles.open(cleanup_path, "wb") as temp_file:
                    await temp_file.write(resp.content)
            actual_path = cleanup_path
        except Exception as exc:  # noqa: BLE001
            logger.error(f"下载远程文件失败: {sanitize_log_message(str(exc))}")
            raise
    else:
        actual_path = Path(file_path)

    try:
        if not actual_path.exists():
            raise FileNotFoundError(f"文件不存在: {actual_path}")

        suffix = actual_path.suffix.lower()

        if suffix == ".pdf":
            return await _process_pdf(actual_path, params, trace_id)
        if suffix in {".txt", ".md"}:
            return await _process_plain_text(actual_path)
        if suffix == ".docx":
            return await _process_docx_with_images(actual_path, params, uploader, trace_id)
        if suffix == ".doc":
            return await _process_doc(actual_path, trace_id)
        if suffix in {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif"}:
            return await _process_image(actual_path, params, trace_id)
        if suffix in {".html", ".htm"}:
            return await _process_html(actual_path)
        if suffix == ".csv":
            return await _process_csv(actual_path)
        if suffix in {".xls", ".xlsx"}:
            return await _process_excel(actual_path)

        raise ValueError(f"不支持的文件类型: {suffix}")
    finally:
        if cleanup_path and cleanup_path.exists():
            cleanup_path.unlink(missing_ok=True)


async def _process_pdf(path: Path, params: dict[str, Any], trace_id: str | None) -> str:
    """
    处理 PDF（仅文本提取，OCR 未内置）

    :param path: PDF 路径
    :param params: 处理参数
    :param trace_id: 链路追踪 ID
    :return:
    """
    logger = get_logger(trace_id)
    enable_ocr = params.get("enable_ocr")
    if enable_ocr and enable_ocr != "disable":
        logger.warning("当前未内置 OCR，PDF 将以文本方式解析")
    reader = PdfReader(path)
    texts = []
    for page in reader.pages:
        texts.append(page.extract_text() or "")
    return "\n\n".join(texts).strip()


async def _process_plain_text(path: Path) -> str:
    async with aiofiles.open(path, encoding="utf-8") as file:
        return await file.read()


async def _process_html(path: Path) -> str:
    async with aiofiles.open(path, encoding="utf-8") as file:
        content = await file.read()
    return md(content, heading_style="ATX")


async def _process_csv(path: Path) -> str:
    df = pd.read_csv(path)
    markdown_parts: list[str] = []
    for _, row in df.iterrows():
        row_df = pd.DataFrame([row], columns=df.columns)
        markdown_parts.append(row_df.to_markdown(index=False))
    return "\n\n".join(markdown_parts).strip()


async def _process_excel(path: Path) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    markdown_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = (
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            )
            value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        columns = _make_unique_columns(data[0] if data else [])
        df_data = data[1:] if len(data) > 1 else []
        df = pd.DataFrame(df_data, columns=columns)

        markdown_parts.append(f"## {sheet_name}")
        table_title = f"{path.stem} - {sheet_name}"
        df.insert(0, "表格标题", table_title)

        chunk_size = 10
        for idx in range(0, len(df), chunk_size):
            chunk_df = df.iloc[idx : idx + chunk_size]
            markdown_parts.append(f"### 数据行 {idx + 1}-{min(idx + chunk_size, len(df))}")
            markdown_parts.append(chunk_df.to_markdown(index=False))

    return "\n\n".join(markdown_parts).strip()


async def _process_docx_with_images(
    path: Path,
    params: dict[str, Any],
    uploader: UploadClient | None,
    trace_id: str | None,
) -> str:
    """
    解析 docx 并上传内嵌图片到 MinIO

    :param path: docx 路径
    :param params: 处理参数
    :param uploader: 上传客户端
    :param trace_id: 链路追踪 ID
    :return:
    """
    if uploader is None:
        raise ValueError("处理 docx 图片需要提供上传客户端")

    db_id = params.get("db_id") or "word-docs"
    logger = get_logger(trace_id)

    with zipfile.ZipFile(path, "r") as zf:
        rels_path = "word/_rels/document.xml.rels"
        rid_to_target: dict[str, str] = {}
        try:
            rels_xml = zf.read(rels_path).decode("utf-8")
            rels_root = _safe_parse_xml(rels_xml)
            for rel in list(rels_root):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                rtype = rel.attrib.get("Type")
                if rid and target and rtype and rtype.endswith("/image"):
                    rid_to_target[rid] = target
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"解析 docx 关系文件失败: {sanitize_log_message(str(exc))}")
            rid_to_target = {}

        document_xml = zf.read("word/document.xml").decode("utf-8")
        ns = {
            "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        root = _safe_parse_xml(document_xml)
        md_lines: list[str] = []

        for paragraph in root.findall(".//w:p", ns):
            texts = [node.text or "" for node in paragraph.findall(".//w:t", ns)]
            para_text = "".join(texts).strip()
            image_urls: list[str] = []

            for blip in paragraph.findall(".//a:blip", ns):
                rid = blip.attrib.get(f"{{{ns['r']}}}embed")
                if not rid:
                    continue
                target = rid_to_target.get(rid)
                if not target:
                    continue
                media_path = target if target.startswith("word/") else f"word/{target}"
                try:
                    data = zf.read(media_path)
                    object_name = f"{db_id}/{uuid4().hex[:8]}/images/{Path(target).name}"
                    content_type = _guess_image_content_type(Path(target).suffix)
                    result = await uploader.upload_bytes(
                        filename=object_name,
                        data=data,
                        content_type=content_type,
                    )
                    image_urls.append(result.url)
                except Exception as exc:  # noqa: BLE001
                    logger.error(f"上传图片失败: {sanitize_log_message(str(exc))}")
                    continue

            line = para_text
            for url in image_urls:
                line = f"{line}\n![image]({url})" if line else f"![image]({url})"
            if line:
                md_lines.append(line)

    return "\n\n".join(md_lines)


async def _process_doc(path: Path, trace_id: str | None) -> str:
    """
    解析 doc（尝试 python-docx，失败提示转为 docx）

    :param path: doc 文件路径
    :param trace_id: 链路追踪 ID
    :return:
    """
    logger = get_logger(trace_id)
    try:
        from docx import Document  # type: ignore
    except Exception as exc:  # noqa: BLE001
        logger.error(f"缺少 python-docx，无法解析 doc: {sanitize_log_message(str(exc))}")
        raise ValueError("解析 doc 需要安装 python-docx，请先安装依赖") from exc

    try:
        doc = Document(path)
        return "\n".join(paragraph.text for paragraph in doc.paragraphs).strip()
    except Exception as exc:  # noqa: BLE001
        logger.error(f"解析 doc 失败: {sanitize_log_message(str(exc))}")
        raise ValueError("无法解析 doc 文件，请转换为 docx 后重试") from exc


async def _process_image(path: Path, params: dict[str, Any], trace_id: str | None) -> str:
    """
    解析图片文本，需外部 OCR 支持

    :param path: 图片路径
    :param params: 处理参数
    :param trace_id: 链路追踪 ID
    :return:
    """
    logger = get_logger(trace_id)
    enable_ocr = params.get("enable_ocr")
    if enable_ocr and enable_ocr != "disable":
        logger.warning("当前未内置 OCR，图片将返回空文本")
        return ""
    raise ValueError("图片解析需要启用 OCR，当前未配置 OCR 引擎")


def _make_unique_columns(columns: Iterable[Any]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for col in columns:
        label = str(col) if col is not None else "Unnamed"
        label = label if label.strip() else "Unnamed"
        if label in seen:
            seen[label] += 1
            unique.append(f"{label}_{seen[label]}")
        else:
            seen[label] = 1
            unique.append(label)
    return unique


def _guess_image_content_type(suffix: str) -> str:
    mapping = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
        ".tif": "image/tiff",
        ".tiff": "image/tiff",
    }
    return mapping.get(suffix.lower(), "image/jpeg")


def _safe_parse_xml(xml_text: str):
    import xml.etree.ElementTree as ET

    return ET.fromstring(xml_text)


def _is_http_url(text: str) -> bool:
    parsed = urlparse(text)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


async def _read_text_file(path: Path, trace_id: str | None) -> str:
    logger = get_logger(trace_id)
    if path.suffix.lower() in {".txt", ".md"}:
        async with aiofiles.open(path, encoding="utf-8") as file:
            return await file.read()

    if path.suffix.lower() == ".pdf":
        return await _process_pdf(path, {}, trace_id)

    logger.warning(f"未针对该类型提供读取实现: {path.suffix}")
    return ""


